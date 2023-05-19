import openpyxl
import requests
import telebot
import time as t
import datetime
import sqlite3
import threading
import schedule
import json
from telebot import types
from datetime import date, time

bot_token = '6224714194:AAERfVMEJn87MunWtYcH8MAtAOkRVelPynE'
bot = telebot.TeleBot(bot_token)

excel_file_path = 'kmb212(1).xlsx'
excel_file_path2 = 'D:\kmb211.xlsx'


conn = sqlite3.connect('kmb2111.db', check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')


conn2 = sqlite3.connect('kmb2112.db', check_same_thread=False)
cursor2 = conn2.cursor()
cursor2.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

conn3 = sqlite3.connect('kmb2211.db', check_same_thread=False)
cursor3 = conn3.cursor()
cursor3.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

conn4 = sqlite3.connect('kmb2212.db', check_same_thread=False)
cursor4 = conn4.cursor()
cursor4.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

conn5 = sqlite3.connect('its2111.db', check_same_thread=False)
cursor5 = conn5.cursor()
cursor5.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')

conn6 = sqlite3.connect('its2112.db', check_same_thread=False)
cursor6 = conn6.cursor()
cursor6.execute('''CREATE TABLE IF NOT EXISTS chat_ids
                (id INTEGER PRIMARY KEY)''')
GET_USER6 = """SELECT chat_ids FROM its2112;"""



@bot.message_handler(commands=['start'])
def start_command(message):

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item1 = types.KeyboardButton("КМБ-21-1")
    item2 = types.KeyboardButton("КМБ-22-1")
    item3 = types.KeyboardButton("ИТС-21-1")
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, "Привет! Нажми на кнопку, чтобы выбрать свою академическую группу", reply_markup=markup)


@bot.message_handler(content_types='text')
def message_reply(message):

    markup2 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup3 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup4 = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item4 = types.KeyboardButton(text="КМБ-21-1(1)")
    item5 = types.KeyboardButton(text="КМБ-21-1(2)")
    markup2.add(item4, item5)

    item6 = types.KeyboardButton("КМБ-22-1(1)")
    item7 = types.KeyboardButton("КМБ-22-1(2)")
    markup3.add(item6, item7)

    item8 = types.KeyboardButton("ИТС-21-1(1)")
    item9 = types.KeyboardButton("ИТС-21-1(2)")
    markup4.add(item8, item9)

    if message.text == "КМБ-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup2)
    if message.text == "КМБ-21-1(1)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        with open("kmb2111.json", "r") as f_o:
            data_from_json=json.load(f_o)
        user_id = message.chat.id

        if str(user_id) not in data_from_json:
            data_from_json.append(user_id)
            # data_from_json.append(" ")

            with open("kmb2111.json", "w") as f_o:
                json.dump(data_from_json, f_o, indent=4, ensure_ascii=False)




        # if weekday_number == 4:
        #     if current_time == "17:41":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)
        #     if current_time == "17:42":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)
        #     if current_time == "17:43":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)


    if message.text == "КМБ-21-1(2)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        # global chat_id
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor2.execute('SELECT * FROM chat_ids WHERE id = ?', (chat_id,))
        result2 = cursor2.fetchone()
        if result2 is None:
            cursor2.execute('INSERT INTO chat_ids VALUES (?)', (chat_id,))
            conn2.commit()
            bot.send_message(message.chat.id, 'ID чата сохранен в базе данных')
        else:
            bot.send_message(message.chat.id, 'ID чата уже существует в базе данных')
        # if weekday_number == 6:
        #     if time == "18:27":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)


    if message.text == "КМБ-22-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup3)


    if message.text == "ИТС-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup4)
    if message.text == "ИТС-21-1(1)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor5.execute('SELECT * FROM chat_ids WHERE id = ?', (chat_id,))
        result5 = cursor5.fetchone()
        if result5 is None:
            cursor5.execute('INSERT INTO chat_ids VALUES (?)', (chat_id,))
            conn5.commit()
    if message.text == "ИТС-21-1(2)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor6.execute('SELECT * FROM chat_ids WHERE id = ?', (chat_id,))

        result6 = cursor6.fetchone()
        if result6 is None:
            cursor6.execute('INSERT INTO chat_ids VALUES (?)', (chat_id,))
            conn6.commit()


def worker():
    while 1:
        today = date.today()
        weekday_number = today.weekday()
        current_time = datetime.datetime.now().strftime('%H:%M:%S')
        print(current_time)
        cou = 1


        cursor2 = conn2.execute("SELECT id FROM chat_ids")
        user_id2 = [row[0] for row in cursor2.fetchall()]
        for id in user_id2:
            wb = openpyxl.load_workbook("kmb212(1).xlsx")
            ws = wb.active
            datafff = ws['B3'].value
            data = ws['A1'].value
            data2 = ws['A2'].value
            data3 = ws['A3'].value
            data4 = ws['A4'].value
            data5 = ws['A5'].value
            data6 = ws['A6'].value
            para2 = ws['B1'].value
            para3 = ws['C1'].value
            para4 = ws['D1'].value
            para5 = ws['E1'].value
            para6 = ws['F1'].value
            para7 = ws['G1'].value
            #print(datafff)
            if cou%2 != 0:
                if weekday_number == 0:
                    if current_time == "07:00:10":
                        bot.send_message(id, ws['A1'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 1:
                    if current_time == "07:00":
                        bot.send_message(id, ws['A2'].value)
                    if current_time == str(para3) or current_time ==str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 2:
                    if current_time == "07:00":
                        bot.send_message(id, ws['A3'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 3:
                    if current_time == "07:00":
                        bot.send_message(id, ws['A4'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 4:
                    if current_time == "07:00":
                        bot.send_message(id, ws['A5'].value)
                    if current_time == str(para3) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 5:
                    if current_time == "07:00":
                        bot.send_message(id, ws['A6'].value)
                    if current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(30)
                if weekday_number == 6:
                    cou=+1
            #else:



thread = threading.Thread(target=worker)
thread.start()
bot.infinity_polling()
